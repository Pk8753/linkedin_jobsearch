from selenium import webdriver
from time import sleep
from bs4 import BeautifulSoup
from openpyxl import load_workbook


class JobUtilites():

    def __init__(self, driver):

        # opening input excel sheet to read data
        self.wb = load_workbook('Driver/input.xlsx')
        self.sheet = self.wb.active

        self.jswb =load_workbook('Result/JobSeeker.xlsx')
        self.page = self.jswb.active

        self.driver = driver

    def jobScrapper(self):
        sleep(2)
        self.driver.find_element_by_xpath("/html/body/div[1]/header/nav/section/section[2]/form/section[1]/input").clear()
        self.driver.find_element_by_xpath("/html/body/div[1]/header/nav/section/section[2]/form/section[1]/input").send_keys(self.sheet['A2'].value)
        self.driver.find_element_by_xpath("/html/body/div[1]/header/nav/section/section[2]/form/section[2]/input").clear()
        self.driver.find_element_by_xpath("/html/body/div[1]/header/nav/section/section[2]/form/section[2]/input").send_keys(self.sheet['B2'].value)
        self.driver.find_element_by_xpath("/html/body/div[1]/header/nav/section/section[2]/form/button").click()
        sleep(2)

    def loadJobIntoExcel(self):

        i = 2
        #scrolling
        ScrollTo = self.driver.find_element_by_xpath('//button[contains(text(),"See more jobs")]')
        sleep(2)
        while (ScrollTo.is_displayed() == False):

            # parsing the visible webpage
            self.pageSource = self.driver.page_source
            self.lxml_soup = BeautifulSoup(self.pageSource, 'lxml')
            self.job_container = self.lxml_soup.find_all('div', class_ = 'base-card base-card--link base-search-card base-search-card--link job-search-card')
            self.driver.execute_script("window.scrollTo(0,document.body.scrollHeight)")

            ScrollTo.location_once_scrolled_into_view
            # searching for all job containers

        sleep(2)
        #
        # self.driver.execute_script("window.scrollTo(0,document.body.scrollHeight)")
        # sleep(2)
        # self.driver.execute_script("window.scrollTo(0,document.body.scrollHeight)")


        print('You are scraping information about {} jobs.'.format(len(self.job_container)))
        for job in self.job_container:

            # job title
            job_titles = job.find("span", class_="screen-reader-text").get_text()
            self.page.cell(row = i, column=1).value= str(job_titles).strip()

            # company name
            company_names = job.find("h4",class_="base-search-card__subtitle").get_text()
            self.page.cell(row=i, column=2).value = str(company_names).strip()

            #job location
            job_locations = job.find("span",class_="job-search-card__location").get_text()
            self.page.cell(row=i, column=3).value = str(job_locations).strip()

            # posting date
            try:
                post_date = job.find("time",class_="job-search-card__listdate--new").get_text()
                self.page.cell(row=i, column=4).value = str(post_date).strip()
            except:
                self.page.cell(row=i, column=4).value = 'Post date is not available'

            #link of the job
            job_link = job.find("a",class_="base-card__full-link" )['href']
            self.page.cell(row=i, column=5).hyperlink = job_link

            i += 1
        self.jswb.save('Result/JobSeeker.xlsx')